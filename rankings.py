import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Color
from openpyxl.formatting.rule import CellIsRule
from io import BytesIO
import os

uploaded_file = st.file_uploader("Choose a file")
if uploaded_file is not None:

    # Load the Excel file
    xls = pd.ExcelFile(uploaded_file)

    # Specify the sheets to combine
    sheets_to_combine = ["且慢投顾", "天天投顾"]

    # Specify the columns of interest
    columns_of_interest = ["组合名称", "成立日期", "成立天数", "成立以来收益", "日张幅", "近一周", "近1月", "近3月", "近6月", "近1年"]

    # Initialize an empty list to store DataFrames
    data_frames = []

    # Loop through the specified sheets
    for sheet in sheets_to_combine:
        if sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)
            
            # Rename columns to standardize across different sheets
            if sheet == "且慢投顾":
                df.rename(columns={
                    '日涨幅': '日张幅',
                    '周涨幅': '近一周',
                    '月涨幅': '近1月',
                    '季涨幅': '近3月',
                    '半年涨幅': '近6月',
                    '年涨幅': '近1年',
                    '成立至今': '成立以来收益'
                }, inplace=True)
            elif sheet == "天天投顾":
                df.rename(columns={
                    '日涨幅': '日张幅',
                    '近一周': '近一周',
                    '近1月': '近1月',
                    '近3月': '近3月',
                    '近6月': '近6月',
                    '近1年': '近1年',
                    '成立来': '成立以来收益'
                }, inplace=True)
            
            # Filter the DataFrame to include only the columns of interest
            df_filtered = df[columns_of_interest].copy()  # Create a copy of the DataFrame slice

            # Calculate the annualized return and add it as a new column using .loc to avoid SettingWithCopyWarning
            df_filtered.loc[:, '年化收益'] = ((df_filtered['成立以来收益'] / df_filtered['成立天数']) * 365).round(2)
            
            data_frames.append(df_filtered)
        else:
            print(f"Sheet {sheet} not found in the Excel file.")

    # Concatenate all the DataFrames into a single DataFrame
    combined_df = pd.concat(data_frames)

    # Load the 关注的策略 sheet
    if "关注的策略" in xls.sheet_names:
        strategy_df = pd.read_excel(xls, sheet_name="关注的策略")
        # Adjust the column name to match the actual name in the sheet
        strategy_df.rename(columns={'策略名称': '组合名称'}, inplace=True)
        strategy_df = strategy_df[['组合名称', '博主名称']]
    else:
        print("Sheet 关注的策略 not found in the Excel file.")

    # Merge the strategy information with the combined DataFrame
    if 'strategy_df' in locals():
        combined_df = pd.merge(combined_df, strategy_df, on='组合名称', how='left')
        # Drop rows where '博主名称' is NaN
        combined_df.dropna(subset=['博主名称'], inplace=True)
        # Move '博主名称' to the first column
        cols = ['博主名称'] + [col for col in combined_df.columns if col != '博主名称']
        combined_df = combined_df[cols]
    else:
        print("Strategy DataFrame not available for merging.")

    # Move '年化收益' to be after '成立以来收益'
    cols = combined_df.columns.tolist()
    # Remove '年化收益' from its current position
    cols.remove('年化收益')
    # Find the index of '成立以来收益'
    index = cols.index('成立以来收益')
    # Insert '年化收益' right after '成立以来收益'
    cols.insert(index + 1, '年化收益')
    # Reorder the DataFrame
    combined_df = combined_df[cols]

    # Convert '成立日期' to datetime and format it
    combined_df['成立日期'] = pd.to_datetime(combined_df['成立日期']).dt.strftime('%Y/%m/%d')

    # Save the updated combined DataFrame to a new Excel file
    output_filename = "output.xlsx"
    output_path = os.path.join(os.getcwd(), output_filename)
    combined_df.to_excel(output_path, index=False)

    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # Apply styles to the first row (header)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='E0E4BC', end_color='E0E4BC', fill_type='solid')

    for cell in ws[1]:  # Row 1 is the header row
        cell.font = header_font
        cell.fill = header_fill

    # Set the background color for the columns '博主名称' and '组合名称', excluding the header row
    column_fill = PatternFill(start_color='DADADA', end_color='DADADA', fill_type='solid')
    columns_of_interest = ['博主名称', '组合名称']

    # Find the columns dynamically
    for col_obj in ws[1]:
        if col_obj.value in columns_of_interest:
            col_letter = col_obj.column_letter
            for row in range(2, ws.max_row + 1):  # Start from row 2 to skip the header
                ws[col_letter + str(row)].fill = column_fill

    # Define the new color fill for specified columns
    new_column_fill = PatternFill(start_color='DBEFF4', end_color='DBEFF4', fill_type='solid')
    columns_to_color = ['成立日期', '成立天数', '成立以来收益']

    # Find the columns dynamically and apply the new color
    for col_obj in ws[1]:  # Loop through the first row which contains the column headers
        if col_obj.value in columns_to_color:
            col_letter = col_obj.column_letter
            for row in range(2, ws.max_row + 1):  # Start from row 2 to skip the header
                ws[col_letter + str(row)].fill = new_column_fill

    new_column_fill = PatternFill(start_color='fff3cd', end_color='fff3cd', fill_type='solid')
    columns_to_color = ["近一周", "近1月", "近3月", "近6月", "近1年"]

    # Find the columns dynamically and apply the new color
    for col_obj in ws[1]:  # Loop through the first row which contains the column headers
        if col_obj.value in columns_to_color:
            col_letter = col_obj.column_letter
            for row in range(2, ws.max_row + 1):  # Start from row 2 to skip the header
                ws[col_letter + str(row)].fill = new_column_fill


    # Define the new color fill for the '年化收益' column
    annual_return_fill = PatternFill(start_color='2596BE', end_color='2596BE', fill_type='solid')

    # Find the '年化收益' column dynamically and apply the new color
    for col_obj in ws[1]:  # Loop through the first row which contains the column headers
        if col_obj.value == '年化收益':
            col_letter = col_obj.column_letter
            for row in range(2, ws.max_row + 1):  # Start from row 2 to skip the header
                ws[col_letter + str(row)].fill = annual_return_fill

    # Define the new color fill for the '年化收益' column
    daily_return_fill = PatternFill(start_color='fac190', end_color='fac190', fill_type='solid')

    # Find the '年化收益' column dynamically and apply the new color
    for col_obj in ws[1]:  # Loop through the first row which contains the column headers
        if col_obj.value == '日张幅':
            col_letter = col_obj.column_letter
            for row in range(2, ws.max_row + 1):  # Start from row 2 to skip the header
                ws[col_letter + str(row)].fill = daily_return_fill

    # Define the columns to apply conditional formatting
    columns_to_format = ['成立以来收益', '年化收益', '日张幅', '近一周', '近1月', '近3月', '近6月', '近1年']

    # Apply conditional formatting to the specified columns
    for col_obj in ws[1]:  # Loop through the first row which contains the column headers
        if col_obj.value in columns_to_format:
            col_letter = col_obj.column_letter
            # Define the range to apply formatting
            cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"
            
            # Apply red font color if the value is positive
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator='greaterThan', formula=['0'], font=Font(color='CF202B'))
            )
            
            # Apply green font color if the value is negative
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator='lessThan', formula=['0'], font=Font(color='79AC73'))
            )

    # Save the changes to the workbook
    wb.save(output_path)
    output_ready = 1 
    # Load the workbook again to create a BytesIO object for downloading
    wb = openpyxl.load_workbook(output_path)
    with BytesIO() as virtual_workbook:
        wb.save(virtual_workbook)
        virtual_workbook.seek(0)  # Go to the beginning of the BytesIO object after saving
        if output_ready == 1:  # Assuming output_file is a condition to check if the file should be offered for download
            st.download_button(
                label="Download data as xlsx",
                data=virtual_workbook,
                file_name="output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )



